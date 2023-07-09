#各種import
import json
import jwt
import requests
import time
import hmac
import hashlib
import time
import websocket
import datetime
import openpyxl as excel

# automail
import pickle
import os.path
from googleapiclient.discovery import build
from google_auth_oauthlib.flow import InstalledAppFlow
from google.auth.transport.requests import Request
import base64
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication
from email.mime.multipart import MIMEMultipart
from apiclient import errors
from os.path import basename

# coincheck側 board
# coincheck 買計算
def c_b_buy():
    global cbd
    global c_buy_hiyou
    URL = 'https://coincheck.com/api/order_books'
    cbd = requests.get(URL).json() 
    x=0
    m=0.005
    c_buy_hiyou=0
    for i in range(100):
        x+=float(cbd["asks"][i][1])
        if x>=0.005:
            c_buy_hiyou+=m*float(cbd["asks"][i][0])
            break
        else:
            m-=float(cbd["asks"][i][1])
            c_buy_hiyou+=float(cbd["asks"][i][1])*float(cbd["asks"][i][0])
            
# coincheck 売計算
def c_b_sell():
    global cbd
    global c_sell_hiyou
    URL = 'https://coincheck.com/api/order_books'
    cbd = requests.get(URL).json() 
    x=0
    m=0.005
    c_sell_hiyou=0
    for i in range(100):
        x+=float(cbd["bids"][i][1])
        if x>=0.005:
            c_sell_hiyou+=m*float(cbd["bids"][i][0])
            break
        else:
            m-=float(cbd["bids"][i][1])
            c_sell_hiyou+=float(cbd["bids"][i][1])*float(cbd["bids"][i][0])


# GMO 側 ボード購入
# GMO 買計算
def g_b_buy():
    global gbd
    global g_buy_hiyou
    endPoint = 'https://api.coin.z.com/public/v1/orderbooks?symbol=BTC'
    gbd = requests.get(endPoint).json()
    x=0
    m=0.005
    g_buy_hiyou=0
    for i in range(100):
        x+=float(gbd["data"]["asks"][i]["size"])
        if x>=0.005:
            g_buy_hiyou+=m*float(gbd["data"]["asks"][i]["price"])
            break
        else:
            m-=float(gbd["data"]["asks"][i]["size"])
            g_buy_hiyou+=float(gbd["data"]["asks"][i]["size"])*float(gbd["data"]["asks"][i]["price"])

# GMO 売計算
def g_b_sell():
    global gbd
    global g_sell_hiyou
    endPoint = 'https://api.coin.z.com/public/v1/orderbooks?symbol=BTC'
    gbd = requests.get(endPoint).json()
    x=0
    m=0.005
    g_sell_hiyou=0
    for i in range(100):
        x+=float(gbd["data"]["bids"][i]["size"])
        if x>=0.005:
            g_sell_hiyou+=m*float(gbd["data"]["bids"][i]["price"])
            break
        else:
            m-=float(gbd["data"]["bids"][i]["size"])
            g_sell_hiyou+=float(gbd["data"]["bids"][i]["size"])*float(gbd["data"]["bids"][i]["price"])

# liquid 買計算
def l_b_buy():
    global lbd
    global l_buy_hiyou
    URL = 'https://api.liquid.com/products/5/price_levels'
    lbd = requests.get(URL).json()
    x=0
    m=0.005
    l_buy_hiyou=0
    for i in range(100):
        x+=float(lbd["sell_price_levels"][i][1])
        if x>=0.005:
            l_buy_hiyou+=m*float(lbd["sell_price_levels"][i][0])
            break
        else:
            m-=float(lbd["sell_price_levels"][i][1])
            l_buy_hiyou+=float(lbd["sell_price_levels"][i][1])*float(lbd["sell_price_levels"][i][0])

# liquid 売計算
def l_b_sell():
    global lbd
    global l_sell_hiyou
    URL = 'https://api.liquid.com/products/5/price_levels'
    lbd = requests.get(URL).json() 
    x=0
    m=0.005
    l_sell_hiyou=0
    for i in range(100):
        x+=float(lbd["buy_price_levels"][i][1])
        if x>=0.005:
            l_sell_hiyou+=m*float(lbd["buy_price_levels"][i][0])
            break
        else:
            m-=float(lbd["buy_price_levels"][i][1])
            l_sell_hiyou+=float(lbd["buy_price_levels"][i][1])*float(lbd["buy_price_levels"][i][0])


def file_name(name):
    global w_name
    global file
    # Excelへの書き込み
    year=datetime.date.today().year
    mon=datetime.date.today().month
    if int(mon)<10:
        mon="0"+str(mon)
    date=datetime.date.today().day
    if int(date)<10:
        date="0"+str(date)
    hour=datetime.datetime.now().hour
    if hour<10:
        hour="0"+str(hour)
    minute=datetime.datetime.now().minute
    if minute<10:
        minute="0"+str(minute)
    sec=datetime.datetime.now().second
    if sec<10:
        sec="0"+str(sec)
    if name=="GMO":
        w_name=f"./datas/GMO/GMO_{str(year)+str(mon)+str(date)}_boards.xlsx"
    elif name=="coincheck":
        w_name=f"./datas/coincheck/coincheck_{str(year)+str(mon)+str(date)}_boards.xlsx"
    elif name=="liquid":
        w_name=f"./datas/liquid/liquid_{str(year)+str(mon)+str(date)}_boards.xlsx"
    else:
        print("file_name()エラー：ネーミングミス?")
    if int(mon)<10:
        mon="0"+str(mon)

def g_write(file,s):
    
    try:
        book=excel.load_workbook(file)
    except FileNotFoundError:
        book = excel.Workbook()
    sheet = book.active
    m_sheet=sheet.max_row

    # 入力
    sheet["A1"]="bank_name"
    sheet["B1"]="time"
    sheet["C1"]="side"
    sheet["D1"]="price"
    sheet["E1"]="Day_of_the_week"

    # 書込み
    sheet["A"+str(m_sheet+1)]="GMO"
    sheet["B"+str(m_sheet+1)]=str(year)+"/"+str(mon)+"/"+str(date)+" "+str(hour)+":"+str(minute)+":"+str(sec)
    if s == "buy":
        sheet["C"+str(m_sheet+1)]="buy"
        sheet["D"+str(m_sheet+1)]=g_buy_hiyou
    else:
        sheet["C"+str(m_sheet+1)]="sell"
        sheet["D"+str(m_sheet+1)]=g_sell_hiyou
    sheet["E"+str(m_sheet+1)]=datetime.date.today().strftime('%A')

    # 保存
    book.save(file)

def c_write(file,s):

    try:
        book=excel.load_workbook(file)
    except FileNotFoundError:
        book = excel.Workbook()
    sheet = book.active
    m_sheet=sheet.max_row

    # 入力
    sheet["A1"]="bank_name"
    sheet["B1"]="time"
    sheet["C1"]="side"
    sheet["D1"]="price"
    sheet["E1"]="Day_of_the_week"

    # 書込み
    sheet["A"+str(m_sheet+1)]="coincheck"
    sheet["B"+str(m_sheet+1)]=str(year)+"/"+str(mon)+"/"+str(date)+" "+str(hour)+":"+str(minute)+":"+str(sec)
    if s == "buy":
        sheet["C"+str(m_sheet+1)]="buy"
        sheet["D"+str(m_sheet+1)]=c_buy_hiyou
    else:
        sheet["C"+str(m_sheet+1)]="sell"
        sheet["D"+str(m_sheet+1)]=c_sell_hiyou
    sheet["E"+str(m_sheet+1)]=datetime.date.today().strftime('%A')

    # 保存
    book.save(file)

def l_write(file,s):

    try:
        book=excel.load_workbook(file)
    except FileNotFoundError:
        book = excel.Workbook()
    sheet = book.active
    m_sheet=sheet.max_row

    # 入力
    sheet["A1"]="bank_name"
    sheet["B1"]="time"
    sheet["C1"]="side"
    sheet["D1"]="price"
    sheet["E1"]="Day_of_the_week"

    # 書込み
    sheet["A"+str(m_sheet+1)]="coincheck"
    sheet["B"+str(m_sheet+1)]=str(year)+"/"+str(mon)+"/"+str(date)+" "+str(hour)+":"+str(minute)+":"+str(sec)
    if s == "buy":
        sheet["C"+str(m_sheet+1)]="buy"
        sheet["D"+str(m_sheet+1)]=c_buy_hiyou
    else:
        sheet["C"+str(m_sheet+1)]="sell"
        sheet["D"+str(m_sheet+1)]=c_sell_hiyou
    sheet["E"+str(m_sheet+1)]=datetime.date.today().strftime('%A')

    # 保存
    book.save(file)

def diff_write():
        
    # file名
    file=f'./datas/all/all_diff_{str(year)+str(mon)+str(date)}.xlsx'
        
    try:
        book=excel.load_workbook(file)
    except FileNotFoundError:
        book = excel.Workbook()
    sheet = book.active
    m_sheet=sheet.max_row

    # 入力
    sheet["A1"]="bank_name"
    sheet["B1"]="time"
    sheet["C1"]="price"
    sheet["D1"]="Day_of_the_week"
   
    # 書込み liquid-coincheck
    sheet["A"+str(m_sheet+1)]="liquid-coincheck"
    sheet["B"+str(m_sheet+1)]=str(year)+"/"+str(mon)+"/"+str(date)+" "+str(hour)+":"+str(minute)+":"+str(sec)
    sheet["C"+str(m_sheet+1)]=l_sell_hiyou-c_buy_hiyou
    sheet["D"+str(m_sheet+1)]=datetime.date.today().strftime('%A')
    
    # 書込み liquid-GMO
    sheet["A"+str(m_sheet+2)]="liquid-GMO"
    sheet["B"+str(m_sheet+2)]=str(year)+"/"+str(mon)+"/"+str(date)+" "+str(hour)+":"+str(minute)+":"+str(sec)
    sheet["C"+str(m_sheet+2)]=l_sell_hiyou-g_buy_hiyou
    sheet["D"+str(m_sheet+2)]=datetime.date.today().strftime('%A')
    
    # 書込み GMO-coincheck
    sheet["A"+str(m_sheet+3)]="GMO-coincheck"
    sheet["B"+str(m_sheet+3)]=str(year)+"/"+str(mon)+"/"+str(date)+" "+str(hour)+":"+str(minute)+":"+str(sec)
    sheet["C"+str(m_sheet+3)]=g_sell_hiyou-c_buy_hiyou
    sheet["D"+str(m_sheet+3)]=datetime.date.today().strftime('%A')
    
    # 書込み GMO-liquid
    sheet["A"+str(m_sheet+4)]="GMO-liquid"
    sheet["B"+str(m_sheet+4)]=str(year)+"/"+str(mon)+"/"+str(date)+" "+str(hour)+":"+str(minute)+":"+str(sec)
    sheet["C"+str(m_sheet+4)]=g_sell_hiyou-l_buy_hiyou
    sheet["D"+str(m_sheet+4)]=datetime.date.today().strftime('%A') 
    
    # 書込み coincheck-liquid
    sheet["A"+str(m_sheet+5)]="coincheck-liquid"
    sheet["B"+str(m_sheet+5)]=str(year)+"/"+str(mon)+"/"+str(date)+" "+str(hour)+":"+str(minute)+":"+str(sec)
    sheet["C"+str(m_sheet+5)]=c_sell_hiyou-l_buy_hiyou
    sheet["D"+str(m_sheet+5)]=datetime.date.today().strftime('%A')
    
    # 書込み coincheck-GMO
    sheet["A"+str(m_sheet+6)]="coincheck-GMO"
    sheet["B"+str(m_sheet+6)]=str(year)+"/"+str(mon)+"/"+str(date)+" "+str(hour)+":"+str(minute)+":"+str(sec)
    sheet["C"+str(m_sheet+6)]=c_sell_hiyou-g_buy_hiyou
    sheet["D"+str(m_sheet+6)]=datetime.date.today().strftime('%A')

    # 保存
    book.save(file)

def g_all():
    c_b_buy()
    c_b_sell()
    g_b_buy()
    g_b_sell()
    l_b_buy()
    l_b_sell() 
    
def write_all():
    # GMO
    file_name("GMO")
    g_write(w_name,"buy")
    g_write(w_name,"sell")
    # coincheck
    file_name("coincheck")
    c_write(w_name,"buy")
    c_write(w_name,"sell")
    # liquid
    file_name("liquid")
    l_write(w_name,"buy")
    l_write(w_name,"sell")
    # all
    diff_write()

# 取引データ mainループ
def main():
    global year
    global mon
    global date
    global hour
    global minute
    global sec
    # 日付時間の取得
    year=datetime.date.today().year
    mon=datetime.date.today().month
    if int(mon)<10:
        mon="0"+str(mon)
    date=datetime.date.today().day
    if int(date)<10:
        date="0"+str(date)
    hour=datetime.datetime.now().hour
    if hour<10:
        hour="0"+str(hour)
    minute=datetime.datetime.now().minute
    if minute<10:
        minute="0"+str(minute)
    sec=datetime.datetime.now().second
    if sec<10:
        sec="0"+str(sec)

    try:
        g_all()
        write_all()
    except:
        print("実行エラー")

