#各種import
import json
import jwt
import requests
import time
import hmac
import hashlib
import time
import websocket
import datetime
import openpyxl as excel

# automail
import pickle
import os.path
from googleapiclient.discovery import build
from google_auth_oauthlib.flow import InstalledAppFlow
from google.auth.transport.requests import Request
import base64
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication
from email.mime.multipart import MIMEMultipart
from apiclient import errors
from os.path import basename

#初期設定値取得
import auth
s_lis=auth.auth()

# Coincheck キーを定義
access_key = s_lis[0]
secret_key = s_lis[1]


# GMO キーを定義
apiKey    = s_lis[2]
secretKey = s_lis[3]

# Liquidキーを定義
secret = s_lis[4]
token = s_lis[5]

# 閾値
thr= s_lis[6]
mergin= s_lis[7]

# 危険 コインチェック買＆売
def post(params):
    global c_res
    # リクエストURL
    url='https://coincheck.com/api/exchange/orders'        
    # 取引のパラメータ
    params = json.dumps(params)
    # nonce をインクリする(unix timeで)
    nonce = str(int(time.time()))
    # signature を定義
    message = nonce + url + params
    signature = hmac.new(
        bytes(secret_key.encode('ascii')),
        bytes(message.encode('ascii')),
        hashlib.sha256
    ).hexdigest()
    # requestに必要な headers を設定。 （最後の超重要は実験的にコメントアウト中）
    headers = {
            'ACCESS-KEY': access_key,
            'ACCESS-NONCE': nonce,
            'ACCESS-SIGNATURE': signature,
            'Content-Type': 'application/json' # 超重要。
    }
    c_res = requests.post(
            url,
            data=params,
            headers=headers
            ).json()

# 危険 GMO 取引
def g_post(st):
    global g_res
    timestamp = '{0}000'.format(int(time.mktime(datetime.datetime.now().timetuple())))
    method    = 'POST'
    endPoint  = 'https://api.coin.z.com/private'
    path      = '/v1/order'
    reqBody = {
        "symbol": "BTC",
        "side": st,
        "executionType": "MARKET",
        "size": "0.005"
    }

    text = timestamp + method + path + json.dumps(reqBody)
    sign = hmac.new(bytes(secretKey.encode('ascii')), bytes(text.encode('ascii')), hashlib.sha256).hexdigest()

    headers = {
        "API-KEY": apiKey,
        "API-TIMESTAMP": timestamp,
        "API-SIGN": sign
    }
    g_res = requests.post(endPoint + path, headers=headers, data=json.dumps(reqBody))


# 資産確認（coincheck）
def c_get():
    global c_res
    global c_yen
    global c_bit
    
    # リクエストURL
    url='https://coincheck.com/api/accounts/balance'
    
    # 毎回インクリする nonce をunixにする
    nonce = str(int(time.time()))
    
    # nonceも使って signature を定義
    message = nonce + url
    signature = hmac.new(
                bytes(secret_key.encode('ascii')),
                bytes(message.encode('ascii')),
                hashlib.sha256
            ).hexdigest()

    # requestに必要な headers を設定。 （最後の超重要は実験的にコメントアウト中）
    headers = {
            'ACCESS-KEY': access_key,
            'ACCESS-NONCE': nonce,
            'ACCESS-SIGNATURE': signature
            #'Content-Type': 'application/json' # 超重要。
    }
    c_res = requests.get(url,headers=headers).json()
    c_yen = float(c_res['jpy'])
    c_bit = float(c_res['btc'])


# 資産確認 (GMO)
def g_get():
    global g_res
    global g_yen
    global g_bit
    timestamp = '{0}000'.format(int(time.mktime(datetime.datetime.now().timetuple())))
    method    = 'GET'
    endPoint  = 'https://api.coin.z.com/private'
    path      = '/v1/account/assets'
    text = timestamp + method + path
    sign = hmac.new(bytes(secretKey.encode('ascii')), bytes(text.encode('ascii')), hashlib.sha256).hexdigest()
    headers = {
        "API-KEY": apiKey,
        "API-TIMESTAMP": timestamp,
        "API-SIGN": sign
    }
    g_res = requests.get(endPoint + path, headers=headers)
    g_yen = float(g_res.json()['data'][0]['amount'])
    g_bit = float(g_res.json()['data'][1]['amount'])

# 取引履歴の確認（coincheck）
def c_hist():
    global ch_res

    # リクエストURL
    url='https://coincheck.com/api/exchange/orders/transactions'
    
    # 毎回インクリする nonce をunixにする
    nonce = str(int(time.time()))
    
    # nonceも使って signature を定義
    message = nonce + url
    signature = hmac.new(
                bytes(secret_key.encode('ascii')),
                bytes(message.encode('ascii')),
                hashlib.sha256
            ).hexdigest()

    # requestに必要な headers を設定。 （最後の超重要は実験的にコメントアウト中）
    headers = {
            'ACCESS-KEY': access_key,
            'ACCESS-NONCE': nonce,
            'ACCESS-SIGNATURE': signature,
            'Content-Type': 'application/json' # 超重要。
            }
    ch_res = requests.get(url,headers=headers).json()

# 取引履歴の確認（GMO）
def gmo_hist():
    global g_hist_res
    timestamp = '{0}000'.format(int(time.mktime(datetime.datetime.now().timetuple())))
    method    = 'GET'
    endPoint  = 'https://api.coin.z.com/private'
    path      = '/v1/latestExecutions'
    text = timestamp + method + path
    sign = hmac.new(bytes(secretKey.encode('ascii')), bytes(text.encode('ascii')), hashlib.sha256).hexdigest()
    parameters = {
        "symbol": "BTC",
        "page": 1,
        "count": 100
    }
    headers = {
        "API-KEY": apiKey,
        "API-TIMESTAMP": timestamp,
        "API-SIGN": sign
    }
    g_hist_res = requests.get(endPoint + path, headers=headers, params=parameters).json()

# coincheck側 board
# coincheck 買計算
def c_b_buy():
    global cbd
    global c_buy_hiyou
    URL = 'https://coincheck.com/api/order_books'
    cbd = requests.get(URL).json() 
    x=0
    m=0.005
    c_buy_hiyou=0
    for i in range(100):
        x+=float(cbd["asks"][i][1])
        if x>=0.005:
            c_buy_hiyou+=m*float(cbd["asks"][i][0])
            break
        else:
            m-=float(cbd["asks"][i][1])
            c_buy_hiyou+=float(cbd["asks"][i][1])*float(cbd["asks"][i][0])
            
# coincheck 売計算
def c_b_sell():
    global cbd
    global c_sell_hiyou
    URL = 'https://coincheck.com/api/order_books'
    cbd = requests.get(URL).json() 
    x=0
    m=0.005
    c_sell_hiyou=0
    for i in range(100):
        x+=float(cbd["bids"][i][1])
        if x>=0.005:
            c_sell_hiyou+=m*float(cbd["bids"][i][0])
            break
        else:
            m-=float(cbd["bids"][i][1])
            c_sell_hiyou+=float(cbd["bids"][i][1])*float(cbd["bids"][i][0])

            
# coincheck 買/売 どちらも計算
def c_b_all():
    global cbd
    global c_buy_hiyou
    global c_sell_hiyou
    URL = 'https://coincheck.com/api/order_books'
    cbd = requests.get(URL).json() 
    # 買ゾーン
    x=0
    m=0.005
    c_buy_hiyou=0
    for i in range(100):
        x+=float(cbd["asks"][i][1])
        if x>=0.005:
            c_buy_hiyou+=m*float(cbd["asks"][i][0])
            break
        else:
            m-=float(cbd["asks"][i][1])
            c_buy_hiyou+=float(cbd["asks"][i][1])*float(cbd["asks"][i][0])
    # 売ゾーン
    x=0
    m=0.005
    c_sell_hiyou=0
    for i in range(100):
        x+=float(cbd["bids"][i][1])
        if x>=0.005:
            c_sell_hiyou+=m*float(cbd["bids"][i][0])
            break
        else:
            m-=float(cbd["bids"][i][1])
            c_sell_hiyou+=float(cbd["bids"][i][1])*float(cbd["bids"][i][0])

# GMO 側 ボード購入
# GMO 買計算
def g_b_buy():
    global gbd
    global g_buy_hiyou
    endPoint = 'https://api.coin.z.com/public/v1/orderbooks?symbol=BTC'
    gbd = requests.get(endPoint).json()
    x=0
    m=0.005
    g_buy_hiyou=0
    for i in range(100):
        x+=float(gbd["data"]["asks"][i]["size"])
        if x>=0.005:
            g_buy_hiyou+=m*float(gbd["data"]["asks"][i]["price"])
            break
        else:
            m-=float(gbd["data"]["asks"][i]["size"])
            g_buy_hiyou+=float(gbd["data"]["asks"][i]["size"])*float(gbd["data"]["asks"][i]["price"])

# GMO 売計算
def g_b_sell():
    global gbd
    global g_sell_hiyou
    endPoint = 'https://api.coin.z.com/public/v1/orderbooks?symbol=BTC'
    gbd = requests.get(endPoint).json()
    x=0
    m=0.005
    g_sell_hiyou=0
    for i in range(100):
        x+=float(gbd["data"]["bids"][i]["size"])
        if x>=0.005:
            g_sell_hiyou+=m*float(gbd["data"]["bids"][i]["price"])
            break
        else:
            m-=float(gbd["data"]["bids"][i]["size"])
            g_sell_hiyou+=float(gbd["data"]["bids"][i]["size"])*float(gbd["data"]["bids"][i]["price"])

            
# coincheck 買/売 どちらも計算
def g_b_all():
    global gbd
    global g_buy_hiyou
    global g_sell_hiyou
    endPoint = 'https://api.coin.z.com/public/v1/orderbooks?symbol=BTC'
    gbd = requests.get(endPoint).json()
    # GMO 買ゾーン
    x=0
    m=0.005
    g_buy_hiyou=0
    for i in range(100):
        x+=float(gbd["data"]["asks"][i]["size"])
        if x>=0.005:
            g_buy_hiyou+=m*float(gbd["data"]["asks"][i]["price"])
            break
        else:
            m-=float(gbd["data"]["asks"][i]["size"])
            g_buy_hiyou+=float(gbd["data"]["asks"][i]["size"])*float(gbd["data"]["asks"][i]["price"])
    # GMO 売ゾーン
    x=0
    m=0.005
    g_sell_hiyou=0
    for i in range(100):
        x+=float(gbd["data"]["bids"][i]["size"])
        if x>=0.005:
            g_sell_hiyou+=m*float(gbd["data"]["bids"][i]["price"])
            break
        else:
            m-=float(gbd["data"]["bids"][i]["size"])
            g_sell_hiyou+=float(gbd["data"]["bids"][i]["size"])*float(gbd["data"]["bids"][i]["price"])    
            
# coincheck の取引履歴表示
def c_ri():
    global c_string
    global c_syupi
    # Nonceは1以上大きくならないといけない
    time.sleep(1)
    c_hist()
    c_syupi=0
    cnt=0
    c_string="コインチェック"
    for i in range(15):
        if ch_res["transactions"][i]["side"]=="sell":
            c_btc=-float(ch_res["transactions"][i]["funds"]["btc"])
        else:
            c_btc=float(ch_res["transactions"][i]["funds"]["btc"])
        cnt+=c_btc
        c_string+=f"coincheckの直近{i+1}回目の取引履歴\n"
        c_string+=f'取引内容：{ch_res["transactions"][i]["side"]}\n'
        c_string+=f'amount ：{c_btc}\n'
        c_string+=f'レート　：{ch_res["transactions"][i]["rate"]}\n'
        c_syupi+=c_btc*float(ch_res["transactions"][i]["rate"])
        print(c_string)
        if cnt > 0.0049:
            c_syupi=c_syupi/cnt*0.005
            break
    if ch_res["transactions"][0]["side"]=="buy":
        print(f"BUY（円換算）：{c_syupi}\n")
        c_string+=f"BUY（円換算）：{c_syupi}\n"
    else:
        print(f"SELL（円換算）：{c_syupi}\n")
        c_string+=f"SELL（円換算）：{c_syupi}\n"
        
        
# gmo の取引履歴表示
def g_ri():
    global g_string
    global g_syupi
    gmo_hist()
    g_syupi=0
    cnt=0
    g_string="GMO\n"
    for i in range(15):
        cnt+=float(g_hist_res["data"]["list"][i]["size"])
        g_string+=f"GMOの直近{i+1}回目の取引履歴\n"
        g_string+=f'取引内容：{g_hist_res["data"]["list"][i]["side"]}\n'
        g_string+=f'amount ：{g_hist_res["data"]["list"][i]["size"]}\n'
        g_string+=f'レート　：{g_hist_res["data"]["list"][i]["price"]}\n'
        print(g_string)
        if g_hist_res["data"]["list"][i]["side"]=="BUY":
            g_syupi+=float(g_hist_res["data"]["list"][i]["size"])*float(g_hist_res["data"]["list"][i]["price"])
        else:
            g_syupi+=float(g_hist_res["data"]["list"][i]["size"])*float(g_hist_res["data"]["list"][i]["price"])
        if cnt > 0.0049:
            if g_hist_res["data"]["list"][i]["side"]=="BUY":
                g_syupi += mergin
                break
            else:
                g_syupi -= mergin
                break
    if g_hist_res["data"]["list"][i]["side"]=="BUY":
        print(f"BUY（円換算）：{g_syupi}\n")
        g_string+=f"BUY（円換算）：{g_syupi}\n"
    else:
        print(f"SELL（円換算）：{g_syupi}\n")
        g_string+=f"SELL（円換算）：{g_syupi}\n"

          
# liquid 買計算
def l_b_buy():
    global lbd
    global l_buy_hiyou
    URL = 'https://api.liquid.com/products/5/price_levels'
    lbd = requests.get(URL).json()
    x=0
    m=0.005
    l_buy_hiyou=0
    for i in range(100):
        x+=float(lbd["sell_price_levels"][i][1])
        if x>=0.005:
            l_buy_hiyou+=m*float(lbd["sell_price_levels"][i][0])
            break
        else:
            m-=float(lbd["sell_price_levels"][i][1])
            l_buy_hiyou+=float(lbd["sell_price_levels"][i][1])*float(lbd["sell_price_levels"][i][0])

# liquid 売計算
def l_b_sell():
    global lbd
    global l_sell_hiyou
    URL = 'https://api.liquid.com/products/5/price_levels'
    lbd = requests.get(URL).json() 
    x=0
    m=0.005
    l_sell_hiyou=0
    for i in range(100):
        x+=float(lbd["buy_price_levels"][i][1])
        if x>=0.005:
            l_sell_hiyou+=m*float(lbd["buy_price_levels"][i][0])
            break
        else:
            m-=float(lbd["buy_price_levels"][i][1])
            l_sell_hiyou+=float(lbd["buy_price_levels"][i][1])*float(lbd["buy_price_levels"][i][0])


# liquid ビットコイン購入
def l_post(params):
    global l_res
    url = 'https://api.liquid.com/orders/'
    payload = {
      "path": '/orders/',
      "nonce": time.mktime(datetime.datetime.now().timetuple()),
      "token_id": token
    }
    signature = jwt.encode(payload, secret, algorithm='HS256')
    headers = {
      'X-Quoine-API-Version': '2',
      'X-Quoine-Auth': signature,
      'Content-Type' : 'application/json'
    }
    data = {
              "order":{
                "order_type": "market",
                "product_id":5,
                "side":params,
                "quantity":0.005
              }
            }
    json_data = json.dumps(data)
    l_res = requests.post(url, headers=headers, data=json_data)


# liquid 資産情報確認
def l_get():
    global l_res
    global l_yen
    global l_bit
    url_str = 'https://api.liquid.com/accounts/balance'
    payload = {
        "path": "/accounts/balance",
        "nonce": time.mktime(datetime.datetime.now().timetuple()),
        "token_id": token
    }
    signature = jwt.encode(
        payload,
        secret,
        algorithm = 'HS256'
    )
    headers = {
        'X-Quoine-API-Version': '2',
        'Content-Type': 'application/json',
        'X-Quoine-Auth': signature
    }

    l_res = (
        requests.get(
            url=url_str,
            headers=headers
        )
    ).json()
    l_yen=float(l_res[0]["balance"])
    l_bit=float(l_res[2]["balance"])

# liquid 取引履歴
def l_hist():
    global l_res
    url_str = 'https://api.liquid.com/executions/me?product_id=5'
    payload = {
        "path": "/executions/me?product_id=5",
        "nonce": time.mktime(datetime.datetime.now().timetuple()),
        "token_id": token
    }
    signature = jwt.encode(
        payload,
        secret,
        algorithm = 'HS256'
    )
    headers = {
        'X-Quoine-API-Version': '2',
        'Content-Type': 'application/json',
        'X-Quoine-Auth': signature
    }

    l_res = (
        requests.get(
            url=url_str,
            headers=headers
        )
    ).json()    
    
# liquid 取引履歴
def l_ri():
    global l_string
    global l_syupi
    l_hist()
    l_syupi=0
    cnt=0
    l_string="liquid\n"
    for i in range(100):
        cnt+=float(l_res["models"][i]["quantity"])
        l_string+=f"liquidの直近{i+1}回目\n"
        l_string+=f'取引内容：{l_res["models"][i]["taker_side"]}\n'
        l_string+=f'amount ：{l_res["models"][i]["quantity"]}\n'
        l_string+=f'レート　：{l_res["models"][i]["price"]}\n'
        print(l_string)
        l_syupi+=float(l_res["models"][i]["quantity"])*float(l_res["models"][i]["price"])
        if cnt >= 0.005:
            break
    if l_res["models"][i]["taker_side"]=="buy":
        print(f"BUY（円換算）：{l_syupi}\n")
        l_string+=f"BUY（円換算）：{l_syupi}\n"
    else:
        print(f"SELL（円換算）：{l_syupi}\n")
        l_string+=f"SELL（円換算）：{l_syupi}\n"

def g_write(file):
    # 日付時間の取得
    year=datetime.date.today().year
    mon=datetime.date.today().month
    if int(mon)<10:
        mon="0"+str(mon)
    date=datetime.date.today().day
    if int(date)<10:
        date="0"+str(date)
    hour=datetime.datetime.now().hour
    if hour<10:
        hour="0"+str(hour)
    minute=datetime.datetime.now().minute
    if minute<10:
        minute="0"+str(minute)
    sec=datetime.datetime.now().second
    if sec<10:
        sec="0"+str(sec)

    # 1秒あけてから書込み
    time.sleep(1)
    # GMO履歴をexcelにoutput
    gmo_hist()
    try:
        book=excel.load_workbook(file)
    except FileNotFoundError:
        book = excel.Workbook()
    sheet = book.active
    m_sheet=sheet.max_row

    # 入力
    sheet["A1"]="executionId"
    sheet["B1"]="bank_name"
    sheet["C1"]="timestamp"
    sheet["D1"]="side"
    sheet["E1"]="price"
    sheet["F1"]="size"
    sheet["G1"]="fee"
    sheet["H1"]="金額（手数料無）"
    sheet["I1"]="金額（手数料有）"
    sheet["J1"]="今日の日付"

    # 取得している最新月
    m=0
    for i in range(50):
        pr=float(g_hist_res["data"]["list"][i]["price"])
        si=float(g_hist_res["data"]["list"][i]["size"])
        m+=si
        sheet["A"+str(m_sheet+1+i)]=g_hist_res["data"]["list"][i]["executionId"]
        sheet["B"+str(m_sheet+1+i)]="GMO"
        sheet["C"+str(m_sheet+1+i)]=g_hist_res["data"]["list"][i]["timestamp"][0:16]
        sheet["D"+str(m_sheet+1+i)]=g_hist_res["data"]["list"][i]["side"]
        sheet["E"+str(m_sheet+1+i)]=g_hist_res["data"]["list"][i]["price"]
        sheet["F"+str(m_sheet+1+i)]=g_hist_res["data"]["list"][i]["size"]
        sheet["G"+str(m_sheet+1+i)]=g_hist_res["data"]["list"][i]["fee"]
        if g_hist_res["data"]["list"][i]["side"] == "SELL":
            sheet["H"+str(m_sheet+1+i)]=pr*si
            sheet["I"+str(m_sheet+1+i)]=pr*si-float(g_hist_res["data"]["list"][i]["fee"])
        else:
            sheet["H"+str(m_sheet+1+i)]=-pr*si
            sheet["I"+str(m_sheet+1+i)]=-pr*si-float(g_hist_res["data"]["list"][i]["fee"])
        sheet["J"+str(m_sheet+1+i)]=str(year)+"/"+str(mon)+"/"+str(date)+" "+str(hour)+":"+str(minute)+":"+str(sec)
        if m>=0.0049999:
            break

    # 保存
    book.save(file)

def c_write(file):
    # 日付時間の取得
    year=datetime.date.today().year
    mon=datetime.date.today().month
    if int(mon)<10:
        mon="0"+str(mon)
    date=datetime.date.today().day
    if int(date)<10:
        date="0"+str(date)
    hour=datetime.datetime.now().hour
    if hour<10:
        hour="0"+str(hour)
    minute=datetime.datetime.now().minute
    if minute<10:
        minute="0"+str(minute)
    sec=datetime.datetime.now().second
    if sec<10:
        sec="0"+str(sec)

    # 1秒あけてから書込み
    time.sleep(1)
    # COINCHECK履歴をexcelにoutput
    c_hist()
    try:
        book=excel.load_workbook(file)
    except FileNotFoundError:
        book = excel.Workbook()
    sheet = book.active
    m_sheet=sheet.max_row

    # sheet["B2"]="こんにちわ"
    sheet["A1"]="executionId"
    sheet["B1"]="bank_name"
    sheet["C1"]="timestamp"
    sheet["D1"]="side"
    sheet["E1"]="price"
    sheet["F1"]="size"
    sheet["G1"]="fee"
    sheet["H1"]="金額（手数料無）"
    sheet["I1"]="金額（手数料有）"
    sheet["J1"]="今日の日付"

    m=0
    for i in range(50):
        pr=float(ch_res["transactions"][i]["rate"])
        if ch_res["transactions"][i]["side"] == "sell":
            si=abs(float(ch_res["transactions"][i]["funds"]["btc"]))
            m+=si
            sheet["H"+str(m_sheet+1+i)]=pr*si
            sheet["I"+str(m_sheet+1+i)]=pr*si
        else:
            si=abs(float(ch_res["transactions"][i]["funds"]["btc"]))
            m+=si
            sheet["H"+str(m_sheet+1+i)]=-pr*si/0.00502*0.005
            sheet["I"+str(m_sheet+1+i)]=-pr*si/0.00502*0.005
        sheet["A"+str(m_sheet+1+i)]=ch_res["transactions"][i]["order_id"]
        sheet["B"+str(m_sheet+1+i)]="CoinCheck"
        sheet["C"+str(m_sheet+1+i)]=ch_res["transactions"][i]["created_at"][0:16]
        sheet["D"+str(m_sheet+1+i)]=ch_res["transactions"][i]["side"]
        sheet["E"+str(m_sheet+1+i)]=ch_res["transactions"][i]["rate"]
        sheet["F"+str(m_sheet+1+i)]=si
        sheet["G"+str(m_sheet+1+i)]=0
        sheet["J"+str(m_sheet+1+i)]=str(year)+"/"+str(mon)+"/"+str(date)+" "+str(hour)+":"+str(minute)+":"+str(sec)
        if m>=0.005:
            break
    # 保存
    book.save(file)

def l_write(file):
    # 日付時間の取得
    year=datetime.date.today().year
    mon=datetime.date.today().month
    if int(mon)<10:
        mon="0"+str(mon)
    date=datetime.date.today().day
    if int(date)<10:
        date="0"+str(date)
    hour=datetime.datetime.now().hour
    if hour<10:
        hour="0"+str(hour)
    minute=datetime.datetime.now().minute
    if minute<10:
        minute="0"+str(minute)
    sec=datetime.datetime.now().second
    if sec<10:
        sec="0"+str(sec)

    # 1秒あけてから書込み
    time.sleep(1)
    # liquid 履歴をexcelにoutput
    l_hist()
    try:
        book=excel.load_workbook(file)
    except FileNotFoundError:
        book = excel.Workbook()
    sheet = book.active
    m_sheet=sheet.max_row

    sheet["A1"]="executionId"
    sheet["B1"]="bank_name"
    sheet["C1"]="timestamp"
    sheet["D1"]="side"
    sheet["E1"]="price"
    sheet["F1"]="size"
    sheet["G1"]="fee"
    sheet["H1"]="金額（手数料無）"
    sheet["I1"]="金額（手数料有）"
    sheet["J1"]="今日の日付"

    m=0
    for i in range(50):
        pr=float(l_res["models"][i]["price"])
        if l_res["models"][i]["taker_side"] == "sell":
            si=abs(float(l_res["models"][i]["quantity"]))
            m+=si
            sheet["H"+str(m_sheet+1+i)]=pr*si
            sheet["I"+str(m_sheet+1+i)]=pr*si
        else:
            si=abs(float(l_res["models"][i]["quantity"]))
            m+=si
            sheet["H"+str(m_sheet+1+i)]=-pr*si
            sheet["I"+str(m_sheet+1+i)]=-pr*si
        sheet["A"+str(m_sheet+1+i)]=l_res["models"][i]["id"]
        sheet["B"+str(m_sheet+1+i)]="liquid"
        sheet["C"+str(m_sheet+1+i)]=datetime.datetime.fromtimestamp(float(l_res["models"][i]["timestamp"])).strftime('%Y-%m-%d %H:%M:%S')
        sheet["D"+str(m_sheet+1+i)]=l_res["models"][i]["taker_side"]
        sheet["E"+str(m_sheet+1+i)]=l_res["models"][i]["price"]
        sheet["F"+str(m_sheet+1+i)]=si
        sheet["G"+str(m_sheet+1+i)]=0
        sheet["J"+str(m_sheet+1+i)]=str(year)+"/"+str(mon)+"/"+str(date)+" "+str(hour)+":"+str(minute)+":"+str(sec)
        if m>=0.0049999:
            break
    # 保存
    book.save(file)

def file_name(name):
    global w_name
    global file
    # Excelへの書き込み
    year=datetime.date.today().year
    mon=datetime.date.today().month
    if int(mon)<10:
        mon="0"+str(mon)
    date=datetime.date.today().day
    if int(date)<10:
        date="0"+str(date)
    hour=datetime.datetime.now().hour
    if hour<10:
        hour="0"+str(hour)
    minute=datetime.datetime.now().minute
    if minute<10:
        minute="0"+str(minute)
    sec=datetime.datetime.now().second
    if sec<10:
        sec="0"+str(sec)
    if name=="GMO":
        w_name=f"./transactions/GMO_{str(year)+str(mon)+str(date)}.xlsx"
    elif name=="coincheck":
        w_name=f"./transactions/coincheck_{str(year)+str(mon)+str(date)}.xlsx"
    elif name=="liquid":
        w_name=f"./transactions/liquid_{str(year)+str(mon)+str(date)}.xlsx"
    else:
        print("file_name()エラー：ネーミングミス?")
    file=f'./transactions/report_{str(year)+str(mon)}.xlsx'

# メール自動送信処理
def auto_send(subject,message_text):
    #  Gmail APIのスコープを設定
    SCOPES = ['https://www.googleapis.com/auth/gmail.send']

    # メールの内容を作成
    msg = MIMEMultipart()

    #  メール本文の作成
    def create_message(sender, to_email, subject, message_text):
        #メール送信先
        msg['to'] = to_email
        # メール送信元
        msg['from'] = sender
        # メールのタイトル(件名)
        msg['subject'] = subject
        # ファイルを添付
        msg.attach(MIMEText(message_text))

        encode_message = base64.urlsafe_b64encode(msg.as_bytes())
        return {'raw': encode_message.decode()}

    #  メール送信の実行
    def send_message(service, user_id, message):
        try:
            message = (service.users().messages().send(userId=user_id, body=message)
                       .execute())
            print('Message Id: %s' % message['id'])
            return message
        except errors.HttpError as error:
            print('An error occurred: %s' % error)

    #  メインとなる処理
    def main():
        #  アクセストークンの取得
        creds = None
        if os.path.exists('token.pickle'):
            with open('token.pickle', 'rb') as token:
                creds = pickle.load(token)
        if not creds or not creds.valid:
            if creds and creds.expired and creds.refresh_token:
                creds.refresh(Request())
            else:
                flow = InstalledAppFlow.from_client_secrets_file(
                    'credentials.json', SCOPES)
                creds = flow.run_local_server()
            with open('token.pickle', 'wb') as token:
                pickle.dump(creds, token)
        service = build('gmail', 'v1', credentials=creds)
        #  メール本文の作成
        sender = 'pazpazmind@gmail.com' # 送信者のアドレス
        to_email = 'bboykou62@yahoo.co.jp' # 受信者のアドレス

        # ファイルを添付
        year=datetime.date.today().year
        mon=datetime.date.today().month
        if int(mon)<10:
            mon="0"+str(mon)
        path = f"./transactions/report_{str(year)+str(mon)}.xlsx"
        with open(path, "rb") as f:
            part = MIMEApplication(
                f.read(),
                Name=basename(path)
            )
        part['Content-Disposition'] = 'attachment; filename="%s"' % basename(path)
        msg.attach(part)


        message = create_message(sender, to_email, subject, message_text)
        #  Gmail APIを呼び出してメール送信
        send_message(service, 'me', message)
    main()

# 送信関数
def sending(a,b,c,d):
    year=str(datetime.date.today().year)
    mon=str(datetime.date.today().month)
    date=str(datetime.date.today().day)
    hour=str(datetime.datetime.now().hour)
    minute=str(datetime.datetime.now().minute)
    sec=str(datetime.datetime.now().second)
    subject=year+"年"+mon+"月"+date+"日"+hour+"時"+minute+"分"+sec+"秒"+" 会員No："+kaiinno
    message_text=f"利益：{c-d}\n"+a+"\n"+b
    auto_send(subject,message_text)

# GMOのステータス
def g_check():
    global g_status
    endPoint = 'https://api.coin.z.com/public'
    path     = '/v1/status'
    response = requests.get(endPoint + path)
    g_status = response.json()["data"]["status"]
    
# liquidのステータス
def l_check():
    global l_status
    global test
    URL = 'https://vvdb046kx1bq.statuspage.io/api/v2/summary.json'
    test = requests.get(URL).json() 
    if test["status"]["description"]!="All Systems Operational":
        l_status="NG"
    else:
        l_status="OPEN"
#     for i in range(len(test["components"])):
#         if test["components"][i]["status"]!="operational":
#             l_status="NG"
#             print(l_status)
#             break
#         else:
#             l_status="OPEN"
#             print(l_status) 


kaiinno="0001"

#自動取引
def main():
    k=0
    ptn=0
    mugen=0
    # レート更新
    c_get()
    g_get()
    l_get()
    c_b_buy()
    c_b_sell()
    g_b_buy()
    g_b_sell()
    l_b_buy()
    l_b_sell()  
    while mugen==0:
        k+=1
        print(k)
        g_check() # GMOのステータス取得
        l_check() # liquidのステータス取得
        try:
            if l_status=="OPEN":
                l_b_sell()
                c_b_buy()
                if (l_sell_hiyou>=c_buy_hiyou+thr)and(l_bit>=0.005)and(c_yen>(c_buy_hiyou*1.02+thr)): # ptn1 liquid売、コインチェック買
                    ptn=1
                    break
            if g_status=="OPEN":
                if l_status=="OPEN":
                    l_b_sell()
                    g_b_buy()
                    if (l_sell_hiyou>=g_buy_hiyou+thr+mergin)and(l_bit>=0.005)and(g_yen>(g_buy_hiyou+thr+mergin)): # ptn2 liquid売、GMO買
                        ptn=2
                        break
            if g_status=="OPEN":
                g_b_sell()
                c_b_buy()
                if (g_sell_hiyou>=c_buy_hiyou+thr+mergin)and(g_bit>=0.005)and(c_yen>(c_buy_hiyou*1.02+thr)): # ptn3 GMO売、coincheck買
                    ptn=3
                    break
                if l_status=="OPEN":
                    g_b_sell()
                    l_b_buy()
                    if (g_sell_hiyou>=l_buy_hiyou+thr+mergin)and(g_bit>=0.005)and(l_yen>(l_buy_hiyou+thr)): # ptn4 GMO売、liquid買
                        ptn=4
                        break
            if l_status=="OPEN":
                c_b_sell()
                l_b_buy()
                if (c_sell_hiyou>=l_buy_hiyou+thr)and(c_bit>=0.005)and(l_yen>(l_buy_hiyou+thr)): # ptn5 coincheck売、liquid買
                    ptn=5
                    break
            if g_status=="OPEN":
                c_b_sell()
                g_b_buy()
                if (c_sell_hiyou>=g_buy_hiyou+thr+mergin)and(c_bit>=0.005)and(g_yen>(g_buy_hiyou+thr+mergin)): # ptn6 coincheck売、GMO買
                    ptn=6
                    break
            print(f"ptn1 liquid売、コインチェック買：{l_sell_hiyou - (c_buy_hiyou+thr)}")
            print(f"ptn2 liquid売、GMO買           ：{l_sell_hiyou - (g_buy_hiyou+thr+mergin)}")
            print(f"ptn3 GMO売、コインチェック買   ：{g_sell_hiyou - (c_buy_hiyou+thr+mergin)}")
            print(f"ptn4 GMO売、liquid買           ：{g_sell_hiyou - (l_buy_hiyou+thr+mergin)}")
            print(f"ptn5 coincheck売、liquid買     ：{c_sell_hiyou - (l_buy_hiyou+thr)}")
            print(f"ptn6 coincheck売、GMO買        ：{c_sell_hiyou - (g_buy_hiyou+thr+mergin)}") 
            print(f"c_bit：{c_bit}")
            print(f"g_bit：{g_bit}")
            print(f"l_bit：{l_bit}\n")
        except:
            print("Error：②何らかのエラーリトライ")
    if ptn==1: # ptn1 liquid売、コインチェック買
        params = {
                    "pair": "btc_jpy", "order_type": "market_buy", "market_buy_amount": round(c_buy_hiyou/0.005*0.00502),
                } # coincheck側買 (金額ではなく量)
        post(params)
        if c_res["success"]==False:
            print("coincheck失敗")
            return
        l_post("sell") # liquid側売
        print(f"ptn1 liquid売、コインチェック買 ：{l_sell_hiyou - (c_buy_hiyou+thr)}")
        time.sleep(1)
        l_ri()
        c_ri()
        print(f"利益：{l_syupi-c_syupi}\n")
        # Excel書込み
        time.sleep(1)
        file_name("liquid")
        l_write(w_name) 
        l_write(file)
        file_name("coincheck")
        c_write(w_name)
        c_write(file)
        sending(l_string,c_string,l_syupi,c_syupi)
    elif ptn==2: # ptn2 liquid売、GMO買
        l_post("sell") # liquid側売
        g_post("BUY")
        print(f"ptn2 liquid売、GMO買            ：{l_sell_hiyou - (g_buy_hiyou+thr+mergin)}")
        time.sleep(1)
        l_ri()
        g_ri()
        print(f"利益：{l_syupi-g_syupi}\n")
        # Excel書込み
        time.sleep(1)
        file_name("liquid")
        l_write(w_name) 
        l_write(file)
        file_name("GMO")
        g_write(w_name)
        g_write(file)
        sending(l_string,g_string,l_syupi,g_syupi)     
    elif ptn==3:# ptn3 GMO売、coincheck買
        params = {     
                    "pair": "btc_jpy", "order_type": "market_buy", "market_buy_amount": round(c_buy_hiyou/0.005*0.00502),
                } # coincheck側買 (金額ではなく量)
        post(params)
        if c_res["success"]==False:
            print("coincheck失敗")
            return
        g_post("SELL") # liquid側売
        print(f"ptn3 GMO売、コインチェック買   ：{g_sell_hiyou - (c_buy_hiyou+thr+mergin)}")
        time.sleep(1)
        g_ri()
        c_ri()
        print(f"利益：{g_syupi-c_syupi}\n")
        # Excel書込み
        time.sleep(1)
        file_name("GMO")
        g_write(w_name) 
        g_write(file)
        file_name("coincheck")
        c_write(w_name)
        c_write(file)
        sending(g_string,c_string,g_syupi,c_syupi)
    elif ptn==4:# ptn4 GMO売、liquid買
        g_post("SELL") # GMO側売
        l_post("buy") # liquid側買
        print(f"ptn4 GMO売、liquid買           ：{g_sell_hiyou - (l_buy_hiyou+thr+mergin)}")
        time.sleep(1)
        g_ri()
        l_ri()
        print(f"利益：{g_syupi-l_syupi}\n") 
        # Excel書込み
        time.sleep(1)
        file_name("liquid")
        l_write(w_name) 
        l_write(file)
        file_name("GMO")
        g_write(w_name)
        g_write(file)
        sending(l_string,g_string,g_syupi,l_syupi)
    elif ptn==5:# ptn5 coincheck売、liquid買
        params = {
            "pair": "btc_jpy",
            "order_type": "market_sell", 
            "amount": 0.005, # 金額ではなく量
        }
        post(params)
        if c_res["success"]==False:
            print("coincheck失敗")
            return
        l_post("buy") # liquid側買
        print(f"ptn5 coincheck売、liquid買     ：{c_sell_hiyou - (l_buy_hiyou+thr)}")
        time.sleep(1)
        c_ri()
        l_ri()
        print(f"利益：{c_syupi-l_syupi}\n")
        # Excel書込み
        time.sleep(1)
        file_name("liquid")
        l_write(w_name) 
        l_write(file)
        file_name("coincheck")
        c_write(w_name)
        c_write(file)
        sending(l_string,c_string,c_syupi,l_syupi)
    elif ptn==6:# ptn6 coincheck売、GMO買
        params = {
            "pair": "btc_jpy",
            "order_type": "market_sell", 
            "amount": 0.005, # 金額ではなく量
        }
        post(params)
        if c_res["success"]==False:
            print("coincheck失敗")
            return
        g_post("BUY") # GMO側買
        print(f"ptn6 coincheck売、GMO買        ：{c_sell_hiyou - (g_buy_hiyou+thr+mergin)}") 
        time.sleep(1)
        c_ri()
        g_ri()
        print(f"利益：{c_syupi-g_syupi}\n")
        # Excel書込み
        time.sleep(1)
        file_name("GMO")
        g_write(w_name) 
        g_write(file)
        file_name("coincheck")
        c_write(w_name)
        c_write(file)
        sending(g_string,c_string,c_syupi,g_syupi)
    else:
        print("謎のエラー")